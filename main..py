from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

root = Tk()
root.title("Airline Reservation System")
root.geometry('900x600+300+200')
root.resizable(False, False)
root.configure(bg="#a2d2ff")

file = pathlib.Path('Backend_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Flight No"
    sheet['B1'] = "Ticket Fare"
    sheet['C1'] = "Origin"
    sheet['D1'] = "Flight Type"
    sheet['E1'] = "Airline Name"
    sheet['F1'] = "Destination"
    sheet['G1'] = "Passenger Name"
    sheet['H1'] = "Passenger Address"
    sheet['I1'] = "Passport Number"
    sheet['J1'] = "Passport Status"

    file.save('Backend_data.xlsx')

def submit():
    flight = flightNo.get()
    ticket = ticketFare.get()
    origin = originValue.get()
    destination = destinationValue.get()
    clas = clas_combobox.get()
    ttkValue = ttk_combobox.get()
    passengerN = addressEntry.get(1.0, END)
    passengerA = passengerAddressEntry.get()
    passportNo = passportNumberEntry.get()
    passportStatus = passportStatusEntry.get()

    file = openpyxl.load_workbook('Backend_data.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=flight)
    sheet.cell(column=2, row=sheet.max_row, value=ticket)
    sheet.cell(column=3, row=sheet.max_row, value=origin)
    sheet.cell(column=4, row=sheet.max_row, value=clas)
    sheet.cell(column=5, row=sheet.max_row, value=ttkValue)
    sheet.cell(column=6, row=sheet.max_row, value=destination)
    sheet.cell(column=7, row=sheet.max_row, value=passengerN)
    sheet.cell(column=8, row=sheet.max_row, value=passengerA)
    sheet.cell(column=9, row=sheet.max_row, value=passportNo)
    sheet.cell(column=10, row=sheet.max_row, value=passportStatus)

    file.save(r'Backend_data.xlsx')

    messagebox.showinfo('info', 'Details Added!')

    flightNo.set('')
    ticketFare.set('')
    originValue.set('')
    destinationValue.set('')
    addressEntry.delete(1.0, END)
    passengerAddressValue.set('')
    passportNumberValue.set('')
    passengerStatusValue.set('')


def clear():
    flightNo.set('')
    originValue.set('')
    destinationValue.set('')
    addressEntry.delete(1.0, END)


# icon
# icon_image=PhotoImage(file="logo.png")
# root.iconphoto(False,icon_image)

# heading
Label(root,
      text="AIRLINE RESERVATION SYSTEM",
      font="helvetica 15",
      bg="#a2d2ff",
      fg="#000").place(x=300, y=5)

Label(root,
      text="---------------------------------------------------------------------------- Reservation Details --------------------------------------------------------------------------------------",

      font="helvetica 15",
      bg="#a2d2ff",
      fg="#fff").place(x=0, y=30)

Label(root,
      text="---------------------------------------------------------------------------- Passenger Details --------------------------------------------------------------------------------------",
      font="helvetica 15",
      bg="#a2d2ff",
      fg="#fff").place(x=0, y=210)
# label

Label(root, text='Flight No:', font=23, bg="#a2d2ff", fg="#000").place(x=30, y=65)
Label(root, text='Ticket Fare:', font=23, bg="#a2d2ff", fg="#000").place(x=360, y=65)
Label(root, text='Origin:', font=23, bg="#a2d2ff", fg="#000").place(x=30, y=120)
Label(root, text='Flight Class:', font=23, bg="#a2d2ff", fg="#000").place(x=360, y=120)
Label(root, text='Flight Name:', font=23, bg="#a2d2ff", fg="#000").place(x=360, y=180)
Label(root, text='Destination:', font=23, bg="#a2d2ff", fg="#000").place(x=30, y=175)
Label(root, text='Passenger Name:', font=23, bg="#a2d2ff", fg="#000").place(x=30, y=250)
Label(root, text='Passenger Address:', font=20, bg="#a2d2ff", fg="#000").place(x=30, y=300)
Label(root, text='Passport No:', font=20, bg="#a2d2ff", fg="#000").place(x=30, y=350)
Label(root, text='Passenger Status:', font=20, bg="#a2d2ff", fg="#000").place(x=30, y=400)

# Entry
flightNo = StringVar()
ticketFare = StringVar()
originValue = StringVar()
destinationValue = StringVar()
ttkValue = StringVar()
passengerAddressValue = StringVar()
passportNumberValue = StringVar()
passengerStatusValue = StringVar()

flightEntry = Entry(root, textvariable=flightNo, width=25, bd=2, font=20)
ticketEntry = Entry(root, textvariable=ticketFare, width=30, bd=2, font=20)
originEntry = Entry(root, textvariable=originValue, width=25, bd=2, font=20)
destinationEntry = Entry(root, textvariable=destinationValue, width=25, bd=2, font=20)
ttkEntry = Entry(root, textvariable=ttkValue, width=25, bd=2, font=20)
passengerAddressEntry = Entry(root, textvariable=passengerAddressValue, width=25, bd=2, font=20)
passportNumberEntry = Entry(root, textvariable=passportNumberValue, width=25, bd=2, font=20)
passportStatusEntry = Entry(root, textvariable=passengerStatusValue, width=25, bd=2, font=20)

# Flighttype
clas_combobox = Combobox(root,
                           values=['First Class', 'Business Class', 'Economy'],
                           font='arial 14',
                           state='r',
                           width=20)
clas_combobox.place(x=440, y=120)
clas_combobox.set('First Class')

ttk_combobox = Combobox(root,
                              values=['Qatar', 'Etihad', 'Lufthansa', 'Air India', 'JAL', 'Delta', 'Southwest'],
                              font='arial 14',
                              state='r',
                              width=20)
ttk_combobox.place(x=440, y=180)
ttk_combobox.set("Qatar")

addressEntry = Text(root, width=40, height=1, bd=2)

flightEntry.place(x=120, y=65)
ticketEntry.place(x=440, y=65)
originEntry.place(x=120, y=120)
destinationEntry.place(x=120, y=175)
addressEntry.place(x=170, y=250)
passengerAddressEntry.place(x=170, y=300)
passportNumberEntry.place(x=170, y=350)
passportStatusEntry.place(x=170, y=400)

Button(root,
       text="Submit",
       bg="#e76f51",
       fg="#000",
       width=10,
       height=2,
       command=submit).place(x=200, y=500)
Button(root,
       text="Clear",
       bg="#e76f51",
       fg="#000",
       width=10,
       height=2,
       command=clear).place(x=340, y=500)
Button(root,
       text="Exit",
       bg="#e76f51",
       fg="#000",
       width=10,
       height=2,
       command=lambda: root.destroy()).place(x=480, y=500)

root.mainloop()
